import asyncio
import html
import os
import re
import zipfile
from datetime import datetime, date, time
from io import BytesIO
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook

# =========================
# STREAMLIT CONFIG
# =========================
st.set_page_config(page_title="Xuất bảng công cho từng nhân viên", layout="wide")
st.title("Xuất bảng công cho từng nhân viên - by Jiangvux")

# Cache Chromium của pyppeteer vào HOME để ổn định trên Streamlit Cloud
os.environ.setdefault("PYPPETEER_HOME", str(Path.home() / ".cache" / "pyppeteer"))

WATERMARK = "Liên hệ Nguyễn Huệ Hr ( 0356 227 868 ) - timvieclam.9phut.com"

CSS = r"""
:root{
  --bg:#ffffff;
  --bar:#3a3a3a;
  --barText:#ffffff;
  --grid:#b7b7b7;
  --head:#f2f2f2;
  --even:#f7f7f7;
  --total:#1b5e20;
  --excelGreen:#217346;
  --danger:#e60023;
  --dangerBg:#ffe6ea;
}
*{box-sizing:border-box;}
html,body{margin:0;padding:0;background:var(--bg);}
body{
  font-family: Arial, Helvetica, "DejaVu Sans", "Liberation Sans", sans-serif;
  color:#111;
}
.wrap{padding:14px 14px 22px;}
.titlebar{
  background:var(--bar);
  color:var(--barText);
  text-align:center;
  font-weight:800;
  padding:12px 14px;
  border-radius:8px;
  font-size:22px;
  letter-spacing:.2px;
  margin:0 0 12px 0;
}
table{
  border-collapse:collapse;
  width:100%;
  table-layout:auto;
  font-size:14px;
}
th, td{
  border:1px solid var(--grid);
  padding:8px 10px;
  text-align:center;
  white-space:nowrap;
}
thead th{background:var(--head);font-weight:800;}
tbody tr:nth-child(even){background:var(--even);}

td.highlight{
  background:var(--excelGreen) !important;
  color:#fff !important;
  font-weight:800;
}

/* Case: có 100% + 130% và tổng < 8 => nền hồng + chữ đỏ + đậm */
td.lowhour{
  background:var(--dangerBg) !important;
  color:var(--danger) !important;
  font-weight:900 !important;
}

/* Case: chỉ có 100% và < 8 => chỉ chữ đỏ + đậm (không tô nền) */
td.lowtext{
  color:var(--danger) !important;
  font-weight:900 !important;
}

tr.total-row td{
  background:var(--total) !important;
  color:#fff !important;
  font-weight:900;
}

.footer-area{
  position:relative;
  margin-top:14px;
  min-height:26px;
}
.stamp{
  position:absolute;
  left:0;
  bottom:0;
  font-size:16px;
  font-weight:800;
  color:#111;
}
.footer{
  text-align:center;
  font-size:18px;
  color:var(--total);
  font-weight:700;
  line-height:26px;
}
"""

# =========================
# HELPERS
# =========================
def safe_sheet_filename(name: str, max_len: int = 90) -> str:
    name = (name or "sheet").strip()
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name)
    name = re.sub(r"\s+", "_", name)
    return name[:max_len] if len(name) > max_len else name

def is_empty_value(v) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s.lower() in ("nan", "none", "nat")

def parse_float(val) -> float:
    if val is None:
        return 0.0
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except:
        return 0.0

def detect_highlight(cell) -> bool:
    """Bắt các ô có tô màu (không tính trắng/đen mặc định)."""
    try:
        fill = cell.fill
        if not fill:
            return False
        fg = fill.fgColor
        if not fg:
            return False
        if getattr(fg, "type", None) == "rgb":
            rgb = (fg.rgb or "").upper()
            return rgb not in ("", "00000000", "FFFFFFFF", "FF000000")
        # Với file dùng theme/pattern, coi là highlight
        if getattr(fill, "patternType", None) or getattr(fill, "fill_type", None):
            return True
        return False
    except:
        return False

def is_date_header(h: str) -> bool:
    h = (h or "").lower().strip()
    return ("ngày" in h) or ("date" in h)

def is_time_header(h: str) -> bool:
    """
    CHỈ coi là cột giờ nếu là các cột vào/ra.
    Tránh nhầm số float (0.5, 1.5...) thành giờ 12:00.
    """
    h = (h or "").lower().strip()
    keys = ["vào", "ra", "gio vao", "gio ra", "giờ vào", "giờ ra", "vao", "ra l", "vào l", "ra lần", "vào lần"]
    return any(k in h for k in keys)

def excel_serial_time_to_hhmm(x: float) -> str:
    total_minutes = int(round(float(x) * 24 * 60))
    h, m = divmod(total_minutes, 60)
    h %= 24
    return f"{h:02d}:{m:02d}"

def format_number(x: float) -> str:
    if abs(x - round(x)) < 1e-9:
        return str(int(round(x)))
    return f"{x:.2f}".rstrip("0").rstrip(".")

def format_cell(value, header_str: str) -> str:
    """
    - Ngày: dd/mm/yyyy
    - Vào/Ra: hh:mm (kể cả float 0..1)
    - Số bình thường (0.5, 1.25...): giữ số, KHÔNG đổi sang 12:00
    """
    h = header_str or ""
    if value is None:
        return ""

    if isinstance(value, datetime):
        if is_date_header(h):
            return value.strftime("%d/%m/%Y")
        if is_time_header(h):
            return value.strftime("%H:%M")
        return value.strftime("%d/%m/%Y")

    if isinstance(value, date) and not isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")

    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"

    if isinstance(value, (float, int)):
        x = float(value)
        if is_time_header(h) and (0 <= x < 1):
            return excel_serial_time_to_hhmm(x)
        return format_number(x)

    return str(value).strip()

def is_col_100(h: str) -> bool:
    """Lương giờ HC tương đương Lương giờ 100%"""
    s = (h or "").lower()
    return ("lương giờ 100%" in s) or ("luong gio 100%" in s) or ("lương giờ hc" in s) or ("luong gio hc" in s) or ("gio 100%" in s)

def is_col_130(h: str) -> bool:
    """Lương giờ ca đêm tương đương Lương giờ 130%"""
    s = (h or "").lower()
    return ("lương giờ 130%" in s) or ("luong gio 130%" in s) or ("lương giờ ca đêm" in s) or ("luong gio ca dem" in s) or ("ca đêm" in s) or ("ca dem" in s) or ("tc 130" in s) or ("tăng ca 130" in s)

def build_html(sheet_name: str, headers: list, rows: list, stamp_text: str) -> str:
    ths = "".join(f"<th>{html.escape(str(h) if h is not None else '')}</th>" for h in headers)

    trs = []
    for r in rows:
        tr_cls = "total-row" if r["is_total"] else ""
        tds = []
        for c in r["cells"]:
            classes = []
            if c.get("highlight"):
                classes.append("highlight")
            if c.get("lowhour"):
                classes.append("lowhour")
            if c.get("lowtext"):
                classes.append("lowtext")
            class_attr = f' class="{" ".join(classes)}"' if classes else ""
            tds.append(f"<td{class_attr}>{c['value_html']}</td>")
        trs.append(f'<tr class="{tr_cls}">{"".join(tds)}</tr>')

    return f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>{CSS}</style>
</head>
<body>
  <div class="wrap">
    <div class="titlebar">{html.escape(sheet_name)}</div>
    <table>
      <thead><tr>{ths}</tr></thead>
      <tbody>
        {''.join(trs)}
      </tbody>
    </table>
    <div class="footer-area">
      <div class="stamp">{html.escape(stamp_text)}</div>
      <div class="footer">{html.escape(WATERMARK)}</div>
    </div>
  </div>
</body>
</html>
"""

def normalize_header(x) -> str:
    return str(x).strip() if x is not None else ""

def unique_preserve_order(items):
    seen = set()
    out = []
    for it in items:
        if it not in seen:
            seen.add(it)
            out.append(it)
    return out

# =========================
# PYPPETEER RENDER (write straight to ZIP + return count)
# =========================
async def render_html_list_to_zip(html_list, zip_file, progress_cb=None):
    """
    Render lần lượt từng HTML -> PNG và ghi trực tiếp vào zip_file.
    Return: số file đã ghi vào ZIP.
    """
    from pyppeteer import launch

    browser = await launch(
        headless=True,
        handleSIGINT=False,
        handleSIGTERM=False,
        handleSIGHUP=False,
        args=[
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-gpu",
            "--disable-dev-shm-usage",
            "--no-zygote",
            "--disable-features=IsolateOrigins,site-per-process",
        ],
    )

    exported = 0
    try:
        page = await browser.newPage()
        await page.setViewport({"width": 1200, "height": 900})

        total = len(html_list)
        for i, (fname, html_str) in enumerate(html_list, start=1):
            if progress_cb:
                progress_cb(i - 1, total, fname)

            await page.setContent(html_str)
            try:
                await page.waitFor(80)
            except:
                pass

            # auto-fit width để không thừa khung trắng
            try:
                dims = await page.evaluate(
                    "() => ({w: document.documentElement.scrollWidth, h: document.documentElement.scrollHeight})"
                )
                w = int(dims.get("w", 1200))
                w = max(980, min(3500, w + 6))
                await page.setViewport({"width": w, "height": 900})
            except:
                pass

            png_bytes = await page.screenshot({"fullPage": True, "type": "png"})
            zip_file.writestr(fname, png_bytes)
            exported += 1

            if progress_cb:
                progress_cb(i, total, fname)

    finally:
        try:
            await browser.close()
        except:
            pass

    return exported

def run_async(coro):
    loop = asyncio.new_event_loop()
    try:
        asyncio.set_event_loop(loop)
        return loop.run_until_complete(coro)
    finally:
        try:
            loop.close()
        except:
            pass

# =========================
# UI: Upload
# =========================
uploaded = st.file_uploader("Chọn file Excel đã tổng hợp (.xlsx)", type=["xlsx"])
if not uploaded:
    st.info("Hãy upload file Excel tổng hợp để bắt đầu.")
    st.stop()

try:
    wb = load_workbook(uploaded, data_only=True)
except Exception as e:
    st.error(f"Không đọc được file Excel. Lỗi: {e}")
    st.stop()

# =========================
# UI: Column selection (NO auto render)
# =========================
all_headers = []
for sn in wb.sheetnames:
    ws = wb[sn]
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not first_row:
        continue
    headers = [normalize_header(h) for h in list(first_row[0])]
    headers = [h for h in headers if h != ""]
    all_headers.extend(headers)

all_headers = unique_preserve_order(all_headers)

st.subheader("Chọn cột muốn in ra ảnh")
st.caption("Tick cột nào thì ảnh sẽ giữ cột đó. Bỏ tick thì loại khỏi ảnh (áp dụng cho tất cả sheet).")

if not all_headers:
    st.warning("Không tìm thấy header (dòng tiêu đề) trong file.")
    st.stop()

if "selected_headers" not in st.session_state:
    st.session_state.selected_headers = set(all_headers)

colA, colB, colC = st.columns([1, 1, 3])
with colA:
    if st.button("✅ Chọn tất cả"):
        st.session_state.selected_headers = set(all_headers)
with colB:
    if st.button("🧹 Bỏ chọn tất cả"):
        st.session_state.selected_headers = set()

search = colC.text_input("🔎 Tìm cột", value="", placeholder="Ví dụ: lương, vào, ra, tăng ca...").strip().lower()
filtered_headers = [h for h in all_headers if (search in h.lower())] if search else all_headers

grid_cols = st.columns(4)
for i, h in enumerate(filtered_headers):
    col = grid_cols[i % 4]
    key = f"col_{all_headers.index(h)}"
    default_checked = h in st.session_state.selected_headers
    checked = col.checkbox(h, value=default_checked, key=key)
    if checked:
        st.session_state.selected_headers.add(h)
    else:
        st.session_state.selected_headers.discard(h)

selected_headers = set(st.session_state.selected_headers)

st.divider()

# =========================
# Render Button
# =========================
btn_col1, btn_col2 = st.columns([1, 3])
with btn_col1:
    render_clicked = st.button("🚀 Xuất File", type="primary", use_container_width=True)
with btn_col2:
    st.caption("Sau khi chọn/bỏ cột xong, bấm nút để bắt đầu render PNG và tải ZIP.")

if not render_clicked:
    st.info("🟦 Chọn cột xong thì bấm **🚀 Xuất File** để bắt đầu.")
    st.stop()

if not selected_headers:
    st.error("Bạn đang bỏ chọn hết cột. Hãy tick ít nhất 1 cột để xuất ảnh.")
    st.stop()

# =========================
# PROCESS SHEETS -> HTML LIST
# =========================
progress = st.progress(0)
status = st.empty()

to_render = []
sheetnames = wb.sheetnames
total_sheets = len(sheetnames)

for idx_sheet, sheet_name in enumerate(sheetnames, start=1):
    status.info(f"Đang chuẩn bị sheet {idx_sheet}/{total_sheets}: **{sheet_name}**")
    ws = wb[sheet_name]

    data = []
    hl = []
    max_cols = 0

    for row in ws.iter_rows():
        row_vals = []
        row_hls = []
        for cell in row:
            row_vals.append("" if cell.value is None else cell.value)
            row_hls.append(detect_highlight(cell))
        max_cols = max(max_cols, len(row_vals))
        data.append(row_vals)
        hl.append(row_hls)

    if len(data) < 2:
        progress.progress(idx_sheet / max(total_sheets, 1))
        continue

    for r in range(len(data)):
        if len(data[r]) < max_cols:
            data[r].extend([""] * (max_cols - len(data[r])))
        if len(hl[r]) < max_cols:
            hl[r].extend([False] * (max_cols - len(hl[r])))

    # keep_cols: cột nào có dữ liệu ở body
    keep_cols = []
    for j in range(max_cols):
        if any(not is_empty_value(data[i][j]) for i in range(1, len(data))):
            keep_cols.append(j)

    if not keep_cols:
        progress.progress(idx_sheet / max(total_sheets, 1))
        continue

    headers_full = [normalize_header(data[0][j]) for j in keep_cols]

    # filter theo checkbox (theo TÊN CỘT)
    keep_cols2 = []
    headers2 = []
    for j, hname in zip(keep_cols, headers_full):
        if hname and (hname in selected_headers):
            keep_cols2.append(j)
            headers2.append(hname)

    if not keep_cols2:
        progress.progress(idx_sheet / max(total_sheets, 1))
        continue

    body_rows_raw = [[data[i][j] for j in keep_cols2] for i in range(1, len(data))]
    hl_rows = [[hl[i][j] for j in keep_cols2] for i in range(1, len(hl))]

    # tìm index cột 100% + 130% trong headers2 (sau khi lọc)
    col_100 = None
    col_130 = None
    for i, h in enumerate(headers2):
        if col_100 is None and is_col_100(h):
            col_100 = i
        if col_130 is None and is_col_130(h):
            col_130 = i

    # add cột STT ở đầu
    headers_out = ["STT"] + headers2

    rows_for_html = []
    stt_counter = 0

    for r_idx, row in enumerate(body_rows_raw):
        first_cell = str(row[0]).strip().lower()
        is_total = first_cell in ("tổng", "tong")

        # STT: chỉ đánh số cho dòng thường, dòng tổng để trống
        stt_val = ""
        if not is_total:
            stt_counter += 1
            stt_val = str(stt_counter)

        # RULE tô đỏ:
        # - Nếu có cả 100% và 130%: chỉ khi có data và (100 + 130) < 8 => tô đỏ cả 2 ô
        # - Nếu chỉ có 100%: chỉ khi có data và 100 < 8 => chữ đỏ + đậm (không nền)
        low_bg_cols = set()    # lowhour
        low_text_cols = set()  # lowtext

        if (not is_total) and (col_100 is not None) and (col_130 is not None):
            raw100 = row[col_100]
            raw130 = row[col_130]
            has_data = (not is_empty_value(raw100)) or (not is_empty_value(raw130))
            if has_data:
                v100 = parse_float(raw100)
                v130 = parse_float(raw130)
                if (v100 + v130) < 8:
                    low_bg_cols.add(col_100)
                    low_bg_cols.add(col_130)

        elif (not is_total) and (col_100 is not None) and (col_130 is None):
            raw100 = row[col_100]
            has_data = not is_empty_value(raw100)
            if has_data:
                v100 = parse_float(raw100)
                if v100 < 8:
                    low_text_cols.add(col_100)

        cells = []

        # STT cell
        cells.append({
            "value_html": html.escape(stt_val),
            "highlight": False,
            "lowhour": False,
            "lowtext": False,
        })

        # data cells
        for c_idx, val in enumerate(row):
            formatted = format_cell(val, headers2[c_idx])
            highlight = bool(hl_rows[r_idx][c_idx]) if r_idx < len(hl_rows) and c_idx < len(hl_rows[r_idx]) else False

            cells.append({
                "value_html": html.escape(formatted),
                "highlight": highlight,
                "lowhour": (c_idx in low_bg_cols),
                "lowtext": (c_idx in low_text_cols),
            })

        rows_for_html.append({"is_total": is_total, "cells": cells})

    stamp_text = safe_sheet_filename(sheet_name)  # đóng dấu góc dưới trái
    html_doc = build_html(sheet_name, headers_out, rows_for_html, stamp_text=stamp_text)
    fname = safe_sheet_filename(sheet_name) + ".png"
    to_render.append((fname, html_doc))

    progress.progress(idx_sheet / max(total_sheets, 1))

status.info("Đang xuất file ảnh..")

if not to_render:
    st.warning("Không có sheet nào có cột được chọn để xuất ảnh.")
    st.stop()

# =========================
# RENDER + ZIP (progress + count)
# =========================
render_bar = st.progress(0)
render_text = st.empty()

def progress_cb(done, total, current_name):
    done = max(0, min(done, total))
    pct = 0 if total == 0 else done / total
    render_bar.progress(pct)
    render_text.info(f"Render {done}/{total}: **{current_name}**")

zip_buf = BytesIO()
exported_count = 0

try:
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as z:
        exported_count = run_async(render_html_list_to_zip(to_render, z, progress_cb=progress_cb))
except Exception as e:
    st.error(
        "Lỗi render Chromium.\n\n"
        f"Chi tiết: {e}\n\n"
        "Nếu browser crash:\n"
        "- Đảm bảo đã commit: runtime.txt (python-3.11), requirements.txt (pyppeteer==1.0.2), packages.txt\n"
        "- Trên Streamlit Cloud bấm Reboot app để rebuild môi trường"
    )
    st.stop()

render_bar.empty()
render_text.empty()

status.success(f"✅ Xong! Đã xuất **{exported_count}** file ảnh và nén ZIP.")
st.download_button(
    "📥 Tải ZIP ảnh PNG",
    data=zip_buf.getvalue(),
    file_name="bang_cong_png.zip",
    mime="application/zip",
)
