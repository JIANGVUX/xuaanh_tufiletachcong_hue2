import asyncio
import hashlib
import html
import os
import re
import sys
import time
import zipfile
from datetime import datetime, date, time as dtime
from io import BytesIO
from pathlib import Path
from tempfile import SpooledTemporaryFile

import streamlit as st
from openpyxl import load_workbook

try:
    import psutil
except Exception:
    psutil = None

try:
    import resource  # Linux/Unix (Streamlit Cloud OK)
except Exception:
    resource = None


# =========================
# STREAMLIT CONFIG
# =========================
st.set_page_config(page_title="Xuất bảng công cho từng nhân viên", layout="wide")
st.title("Xuất bảng công cho từng nhân viên - by Jiangvux")

# Cache Chromium của pyppeteer vào HOME để ổn định trên Streamlit Cloud
os.environ.setdefault("PYPPETEER_HOME", str(Path.home() / ".cache" / "pyppeteer"))

DEFAULT_WATERMARK = "Liên hệ Nguyễn Huệ Hr ( 0356 227 868 ) - timvieclam.9phut.com"
DEFAULT_THRESHOLD = 8.0

DEFAULT_RENAME_RULES = """Lương giờ 100% = Lương giờ ca ngày
Lương giờ 130% = Lương giờ ca đêm
TC 150% = T/ca ngày
TC 200% = T/ca Đêm
TC ngày CN 200% = T/ca giờ ngày CN
TC đêm CN 270% = T/Ca giờ đêm CN
"""


# =========================
# UI: Sidebar config
# =========================
st.sidebar.header("⚙️ Cấu hình xuất ảnh")

device_scale = st.sidebar.slider(
    "Độ nét (deviceScaleFactor) — càng cao càng nét nhưng nặng/chậm",
    min_value=1.0, max_value=3.0, value=1.8, step=0.1
)
auto_fit_width = st.sidebar.checkbox("Auto-fit chiều ngang (giảm thừa khung trắng)", value=True)

min_width = st.sidebar.number_input("Min width (px)", min_value=600, max_value=6000, value=980, step=10)
max_width = st.sidebar.number_input("Max width (px)", min_value=800, max_value=12000, value=3500, step=10)

wait_ms = st.sidebar.slider("Chờ layout sau setContent (ms)", min_value=0, max_value=300, value=60, step=10)

add_stt = st.sidebar.checkbox("Thêm cột STT", value=True)

threshold_hours = st.sidebar.number_input(
    "Ngưỡng giờ chuẩn (>= ngưỡng thì để nguyên, < ngưỡng thì tô đỏ)",
    min_value=0.0, max_value=24.0, value=float(DEFAULT_THRESHOLD), step=0.5
)

st.sidebar.divider()
st.sidebar.subheader("Watermark / Stamp (đóng dấu)")

show_watermark = st.sidebar.checkbox("Hiển thị Watermark (giữa dưới)", value=True)
watermark_text = st.sidebar.text_input("Nội dung Watermark", value=DEFAULT_WATERMARK)
watermark_size = st.sidebar.slider("Cỡ chữ Watermark", 10, 30, 18)

show_stamp = st.sidebar.checkbox("Hiển thị Stamp (góc dưới trái)", value=True)
stamp_mode = st.sidebar.selectbox("Nội dung Stamp", ["Tên sheet", "Tên file ảnh", "Tùy chỉnh"], index=0)
stamp_custom = st.sidebar.text_input("Stamp tùy chỉnh", value="", disabled=(stamp_mode != "Tùy chỉnh"))
stamp_size = st.sidebar.slider("Cỡ chữ Stamp", 10, 26, 16)

st.sidebar.divider()
st.sidebar.subheader("🧩 Đổi tên cột khi xuất ảnh")

rename_enable = st.sidebar.checkbox("Bật đổi tên header khi xuất ảnh", value=True)
rename_rules = st.sidebar.text_area(
    "Nhập mapping dạng: Tên cũ = Tên mới (mỗi dòng 1 mapping)",
    value=DEFAULT_RENAME_RULES,
    height=160,
    disabled=(not rename_enable),
)

st.sidebar.caption("Gợi ý: nếu thấy chậm, giảm deviceScaleFactor xuống 1.4–1.8 và giảm số cột giữ lại.")


# =========================
# RAM UI
# =========================
mem_box = st.sidebar.expander("📈 RAM đang dùng", expanded=True)
mem_placeholder = mem_box.empty()
_mem_peak_mb = 0.0
_t0 = time.perf_counter()

def _read_ram():
    cur_mb = None
    peak_mb = None
    sys_used_pct = None
    sys_avail_mb = None
    sys_total_mb = None

    if psutil is not None:
        try:
            p = psutil.Process(os.getpid())
            cur_mb = p.memory_info().rss / (1024 * 1024)

            vm = psutil.virtual_memory()
            sys_used_pct = float(vm.percent)
            sys_avail_mb = vm.available / (1024 * 1024)
            sys_total_mb = vm.total / (1024 * 1024)
        except Exception:
            pass

    if resource is not None:
        try:
            ru = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
            peak_mb = (ru / (1024 * 1024)) if sys.platform == "darwin" else (ru / 1024.0)
        except Exception:
            pass

    if peak_mb is None:
        peak_mb = cur_mb

    return cur_mb, peak_mb, sys_used_pct, sys_avail_mb, sys_total_mb

def _render_mem_ui(cur_mb, peak_mb, sys_used_pct, sys_avail_mb, sys_total_mb):
    lines = []
    if cur_mb is not None:
        lines.append(f"**RAM process (hiện tại):** `{cur_mb:.0f} MB`")
    else:
        lines.append("**RAM process (hiện tại):** `—` *(cài psutil để đo chuẩn)*")

    if peak_mb is not None:
        lines.append(f"**RAM process (peak):** `{peak_mb:.0f} MB`")

    if sys_used_pct is not None and sys_avail_mb is not None and sys_total_mb is not None:
        lines.append(f"**RAM hệ thống:** `{sys_used_pct:.0f}%` — còn `{sys_avail_mb:.0f}/{sys_total_mb:.0f} MB`")
    else:
        lines.append("**RAM hệ thống:** `—`")

    elapsed = time.perf_counter() - _t0
    lines.append(f"**Thời gian chạy:** `{elapsed:.1f}s`")
    mem_placeholder.markdown("\n\n".join(lines))


# =========================
# CSS (dynamic)
# =========================
def build_css(wm_size: int, sp_size: int) -> str:
    return f"""
:root {{
  --bg:#ffffff;
  --bar:#3a3a3a;
  --barText:#ffffff;
  --grid:#b7b7b7;
  --head:#f2f2f2;
  --even:#f7f7f7;
  --total:#1b5e20;
  --excelGreen:#217346;
  --danger:#e60023;
  --dangerBg:#ffe6ea;
}}
*{{box-sizing:border-box;}}
html,body{{margin:0;padding:0;background:var(--bg);}}
body{{
  font-family: Arial, Helvetica, "DejaVu Sans", "Liberation Sans", sans-serif;
  color:#111;
}}
.wrap{{padding:14px 14px 22px;}}
.titlebar{{
  background:var(--bar);
  color:var(--barText);
  text-align:center;
  font-weight:800;
  padding:12px 14px;
  border-radius:8px;
  font-size:22px;
  letter-spacing:.2px;
  margin:0 0 12px 0;
}}
table{{
  border-collapse:collapse;
  width:100%;
  table-layout:auto;
  font-size:14px;
}}
th,td{{
  border:1px solid var(--grid);
  padding:8px 10px;
  text-align:center;
  white-space:nowrap;
}}
thead th{{background:var(--head);font-weight:800;}}
tbody tr:nth-child(even){{background:var(--even);}}

/* Highlight từ Excel */
td.highlight{{
  background:var(--excelGreen) !important;
  color:#fff !important;
  font-weight:800;
}}

/* Có 100% + 130% và tổng < ngưỡng => nền hồng + chữ đỏ + đậm */
td.lowhour{{
  background:var(--dangerBg) !important;
  color:var(--danger) !important;
  font-weight:900 !important;
}}

/* Chỉ có 100% và < ngưỡng => chữ đỏ + đậm (không tô nền) */
td.lowtext{{
  color:var(--danger) !important;
  font-weight:900 !important;
}}

tr.total-row td{{
  background:var(--total) !important;
  color:#fff !important;
  font-weight:900;
}}

.footer-area{{
  position:relative;
  margin-top:14px;
  min-height:26px;
}}
.stamp{{
  position:absolute;
  left:0;
  bottom:0;
  font-size:{sp_size}px;
  font-weight:800;
  color:#111;
}}
.footer{{
  text-align:center;
  font-size:{wm_size}px;
  color:var(--total);
  font-weight:700;
  line-height:26px;
}}
"""


# =========================
# HELPERS
# =========================
def safe_sheet_filename(name: str, max_len: int = 90) -> str:
    name = (name or "sheet").strip()
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name)
    name = re.sub(r"\s+", "_", name)
    return name[:max_len] if len(name) > max_len else name

def is_empty_value(v) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s.lower() in ("nan", "none", "nat")

def parse_float(val) -> float:
    if val is None:
        return 0.0
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0

def detect_highlight(cell) -> bool:
    try:
        fill = cell.fill
        if not fill:
            return False
        fg = fill.fgColor
        if not fg:
            return False
        if getattr(fg, "type", None) == "rgb":
            rgb = (fg.rgb or "").upper()
            return rgb not in ("", "00000000", "FFFFFFFF", "FF000000")
        if getattr(fill, "patternType", None) or getattr(fill, "fill_type", None):
            return True
        return False
    except Exception:
        return False

def is_date_header(h: str) -> bool:
    h = (h or "").lower().strip()
    return ("ngày" in h) or ("date" in h)

def is_time_header(h: str) -> bool:
    h = (h or "").lower().strip()
    keys = ["vào", "ra", "gio vao", "gio ra", "giờ vào", "giờ ra", "vao", "ra l", "vào l", "ra lần", "vào lần"]
    return any(k in h for k in keys)

def excel_serial_time_to_hhmm(x: float) -> str:
    total_minutes = int(round(float(x) * 24 * 60))
    hh, mm = divmod(total_minutes, 60)
    hh %= 24
    return f"{hh:02d}:{mm:02d}"

def format_number(x: float) -> str:
    if abs(x - round(x)) < 1e-9:
        return str(int(round(x)))
    return f"{x:.2f}".rstrip("0").rstrip(".")

def format_cell(value, header_str: str) -> str:
    h = header_str or ""
    if value is None:
        return ""

    if isinstance(value, datetime):
        if is_date_header(h):
            return value.strftime("%d/%m/%Y")
        if is_time_header(h):
            return value.strftime("%H:%M")
        return value.strftime("%d/%m/%Y")

    if isinstance(value, date) and not isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")

    if isinstance(value, dtime):
        return f"{value.hour:02d}:{value.minute:02d}"

    if isinstance(value, (float, int)):
        x = float(value)
        if is_time_header(h) and (0 <= x < 1):
            return excel_serial_time_to_hhmm(x)
        return format_number(x)

    return str(value).strip()

def is_col_100(h: str) -> bool:
    s = (h or "").lower()
    return ("lương giờ 100%" in s) or ("luong gio 100%" in s) or ("lương giờ hc" in s) or ("luong gio hc" in s) or ("gio 100%" in s)

def is_col_130(h: str) -> bool:
    s = (h or "").lower()
    return ("lương giờ 130%" in s) or ("luong gio 130%" in s) or ("lương giờ ca đêm" in s) or ("luong gio ca dem" in s) or ("ca đêm" in s) or ("ca dem" in s) or ("tc 130" in s) or ("tăng ca 130" in s)

def normalize_header(x) -> str:
    return str(x).strip() if x is not None else ""

def unique_preserve_order(items):
    seen = set()
    out = []
    for it in items:
        if it not in seen:
            seen.add(it)
            out.append(it)
    return out

def is_total_row(row_vals) -> bool:
    for v in row_vals[: min(6, len(row_vals))]:
        if str(v).strip().lower() in ("tổng", "tong"):
            return True
    return False

def is_text_time_like(s: str) -> bool:
    # Chuỗi giờ kiểu 07:30, 19:01...
    if not s:
        return False
    s = s.strip()
    return bool(re.match(r"^\d{1,2}:\d{2}$", s))

def is_numeric_like(v) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float)):
        try:
            return not (v != v)  # not NaN
        except Exception:
            return True
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return False
        if is_text_time_like(s):
            return False
        # số có thể dạng 10,5 hoặc 10.5 hoặc 1,000.5 (thô)
        s2 = s.replace(" ", "")
        s2 = s2.replace(",", ".")
        return bool(re.match(r"^-?\d+(\.\d+)?$", s2))
    return False

def should_sum_by_header(h: str) -> bool:
    """
    Quy tắc chọn cột để cộng:
    - Loại cột ngày/giờ vào-ra
    - Loại cột mô tả: STT, Thứ, Ngày, Mã NV, Họ tên, Bộ phận, Ca...
    - Còn lại nếu dữ liệu là số => cộng
    """
    if not h:
        return True

    hh = h.strip().lower()

    # Loại theo nhận diện ngày/giờ
    if is_date_header(h) or is_time_header(h):
        return False

    # Loại các cột text cố định
    deny_keys = [
        "stt", "thứ", "thu", "ngày", "ngay",
        "mã nv", "ma nv", "họ tên", "ho ten",
        "bộ phận", "bo phan", "ca",
        "ghi chú", "ghi chu", "note", "chú thích", "chu thich"
    ]
    for k in deny_keys:
        if k in hh:
            return False

    return True

def compute_column_sums(body_rows_raw: list, headers2: list) -> tuple:
    """
    Tính tổng theo cột (chỉ các dòng không phải TỔNG).
    Return: (sums, counts)
      - sums[j] = tổng số
      - counts[j] = số lượng ô số đã cộng (để biết cột đó có dữ liệu số hay không)
    """
    ncol = len(headers2)
    sums = [0.0] * ncol
    counts = [0] * ncol

    for row in body_rows_raw:
        if is_total_row(row):
            continue
        for j in range(ncol):
            if not should_sum_by_header(headers2[j]):
                continue
            v = row[j]
            if is_empty_value(v):
                continue

            # loại kiểu datetime/date/time
            if isinstance(v, (datetime, date, dtime)):
                continue

            if isinstance(v, (int, float)):
                num = float(v)
                sums[j] += num
                counts[j] += 1
            elif isinstance(v, str):
                s = v.strip()
                if s == "" or is_text_time_like(s):
                    continue
                if is_numeric_like(s):
                    num = parse_float(s)
                    sums[j] += num
                    counts[j] += 1

    return sums, counts

def fill_total_row_if_missing(body_rows_raw: list, hl_rows: list, headers2: list, force_overwrite: bool = False):
    """
    - Tìm hàng TỔNG nếu có.
    - Tự tính tổng các cột số.
    - Điền vào ô tổng nếu ô đang trống (hoặc ghi đè nếu force_overwrite=True).
    - Nếu không có hàng TỔNG: tự tạo 1 hàng TỔNG ở cuối.
    """
    sums, counts = compute_column_sums(body_rows_raw, headers2)

    total_idx = None
    for i, row in enumerate(body_rows_raw):
        if is_total_row(row):
            total_idx = i
            break

    ncol = len(headers2)

    # Nếu chưa có hàng tổng -> tạo mới
    if total_idx is None:
        new_total = [None] * ncol

        # đặt chữ TỔNG vào 1 cột text hợp lý (ưu tiên cột "Thứ", nếu không có thì cột đầu)
        label_col = 0
        for j, h in enumerate(headers2):
            hs = (h or "").lower()
            if "thứ" in hs or "thu" in hs:
                label_col = j
                break
        new_total[label_col] = "TỔNG"

        # điền tổng
        for j in range(ncol):
            if counts[j] <= 0:
                continue
            if not should_sum_by_header(headers2[j]):
                continue
            val = sums[j]
            new_total[j] = int(round(val)) if abs(val - round(val)) < 1e-9 else val

        body_rows_raw.append(new_total)
        hl_rows.append([False] * ncol)
        return

    # Có hàng tổng -> điền thiếu (hoặc ghi đè)
    total_row = body_rows_raw[total_idx]
    for j in range(ncol):
        if counts[j] <= 0:
            continue
        if not should_sum_by_header(headers2[j]):
            continue

        need_fill = force_overwrite or is_empty_value(total_row[j])
        if not need_fill:
            continue

        val = sums[j]
        total_row[j] = int(round(val)) if abs(val - round(val)) < 1e-9 else val

    body_rows_raw[total_idx] = total_row


def _norm_key(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

def parse_rename_rules(text: str) -> dict:
    """
    Parse mapping lines:
      A = B
      A => B
      A -> B
    Return: dict normalized_old -> new_name (trimmed)
    """
    mp = {}
    if not text:
        return mp
    for raw in str(text).splitlines():
        line = raw.strip()
        if not line:
            continue
        if line.startswith("#") or line.startswith("//"):
            continue

        sep = None
        if "=>" in line:
            sep = "=>"
        elif "->" in line:
            sep = "->"
        elif "=" in line:
            sep = "="

        if not sep:
            continue

        left, right = line.split(sep, 1)
        old = left.strip()
        new = right.strip()
        if not old or not new:
            continue
        mp[_norm_key(old)] = new
    return mp

RENAME_MAP = parse_rename_rules(rename_rules) if rename_enable else {}

def rename_header_for_output(h: str) -> str:
    if not rename_enable or not RENAME_MAP:
        return h
    return RENAME_MAP.get(_norm_key(h), h)

def build_html(sheet_name: str, headers: list, rows: list, stamp_text: str, cfg_css: str,
              watermark: str, show_wm: bool, show_st: bool) -> str:
    ths = "".join(f"<th>{html.escape(str(h) if h is not None else '')}</th>" for h in headers)

    trs = []
    for r in rows:
        tr_cls = "total-row" if r["is_total"] else ""
        tds = []
        for c in r["cells"]:
            classes = []
            if c.get("highlight"):
                classes.append("highlight")
            if c.get("lowhour"):
                classes.append("lowhour")
            if c.get("lowtext"):
                classes.append("lowtext")
            class_attr = f' class="{" ".join(classes)}"' if classes else ""
            tds.append(f"<td{class_attr}>{c['value_html']}</td>")
        trs.append(f'<tr class="{tr_cls}">{"".join(tds)}</tr>')

    stamp_html = f'<div class="stamp">{html.escape(stamp_text)}</div>' if (show_st and stamp_text) else ""
    wm_html = f'<div class="footer">{html.escape(watermark)}</div>' if (show_wm and watermark) else ""

    return f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>{cfg_css}</style>
</head>
<body>
  <div class="wrap">
    <div class="titlebar">{html.escape(sheet_name)}</div>
    <table>
      <thead><tr>{ths}</tr></thead>
      <tbody>
        {''.join(trs)}
      </tbody>
    </table>
    <div class="footer-area">
      {stamp_html}
      {wm_html}
    </div>
  </div>
</body>
</html>
"""


# =========================
# PYPPETEER RENDER
# =========================
async def render_html_list_to_zip(
    html_list,
    zip_file,
    progress_cb=None,
    dsf: float = 1.8,
    auto_fit: bool = True,
    min_w: int = 980,
    max_w: int = 3500,
    wait_ms: int = 60,
):
    from pyppeteer import launch

    browser = await launch(
        headless=True,
        handleSIGINT=False,
        handleSIGTERM=False,
        handleSIGHUP=False,
        args=[
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-gpu",
            "--disable-dev-shm-usage",
            "--no-zygote",
            "--disable-features=IsolateOrigins,site-per-process",
        ],
    )

    exported = 0
    try:
        page = await browser.newPage()
        await page.setViewport({"width": int(min_w), "height": 900, "deviceScaleFactor": float(dsf)})

        total = len(html_list)
        for i, (fname, html_str) in enumerate(html_list, start=1):
            if progress_cb:
                progress_cb(i - 1, total, fname)

            await page.setContent(html_str)

            if wait_ms > 0:
                try:
                    await page.waitFor(int(wait_ms))
                except Exception:
                    pass

            if auto_fit:
                try:
                    dims = await page.evaluate(
                        "() => ({w: document.documentElement.scrollWidth, h: document.documentElement.scrollHeight})"
                    )
                    w = int(dims.get("w", min_w))
                    w = max(int(min_w), min(int(max_w), w + 6))
                    await page.setViewport({"width": w, "height": 900, "deviceScaleFactor": float(dsf)})
                except Exception:
                    pass

            png_bytes = await page.screenshot({"fullPage": True, "type": "png"})
            zip_file.writestr(fname, png_bytes)
            exported += 1

            if progress_cb:
                progress_cb(i, total, fname)

    finally:
        try:
            await browser.close()
        except Exception:
            pass

    return exported

def run_async(coro):
    loop = asyncio.new_event_loop()
    try:
        asyncio.set_event_loop(loop)
        return loop.run_until_complete(coro)
    finally:
        try:
            loop.close()
        except Exception:
            pass


# =========================
# UI: Upload + cache workbook
# =========================
uploaded = st.file_uploader("Chọn file Excel đã tổng hợp (.xlsx)", type=["xlsx"])
if not uploaded:
    st.info("Hãy upload file Excel tổng hợp để bắt đầu.")
    st.stop()

file_bytes = uploaded.getvalue()
file_md5 = hashlib.md5(file_bytes).hexdigest()

if st.session_state.get("file_md5") != file_md5:
    st.session_state.file_md5 = file_md5
    st.session_state.wb = load_workbook(BytesIO(file_bytes), data_only=True)
    st.session_state.all_headers = None
    st.session_state.selected_headers = None
    for k in list(st.session_state.keys()):
        if isinstance(k, str) and k.startswith("hdr__"):
            del st.session_state[k]

wb = st.session_state.wb


# =========================
# UI: Column selection
# =========================
def hdr_key(h: str) -> str:
    hx = hashlib.md5(h.encode("utf-8")).hexdigest()[:10]
    return f"hdr__{hx}"

if st.session_state.get("all_headers") is None:
    all_headers = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if not first_row:
            continue
        headers = [normalize_header(x) for x in list(first_row[0])]
        headers = [h for h in headers if h != ""]
        all_headers.extend(headers)
    st.session_state.all_headers = unique_preserve_order(all_headers)

if st.session_state.get("selected_headers") is None:
    st.session_state.selected_headers = set(st.session_state.all_headers)

all_headers = st.session_state.all_headers

st.subheader("Chọn cột muốn in ra ảnh")
st.caption("Tick cột nào thì ảnh sẽ giữ cột đó. Bỏ tick thì loại khỏi ảnh (áp dụng cho tất cả sheet).")

if not all_headers:
    st.warning("Không tìm thấy header (dòng tiêu đề) trong file.")
    st.stop()

colA, colB, colC = st.columns([1, 1, 3])
with colA:
    if st.button("✅ Chọn tất cả"):
        st.session_state.selected_headers = set(all_headers)
        for h in all_headers:
            st.session_state[hdr_key(h)] = True
with colB:
    if st.button("🧹 Bỏ chọn tất cả"):
        st.session_state.selected_headers = set()
        for h in all_headers:
            st.session_state[hdr_key(h)] = False

search = colC.text_input("🔎 Tìm cột", value="", placeholder="Ví dụ: lương, vào, ra, tăng ca...").strip().lower()
filtered_headers = [h for h in all_headers if (search in h.lower())] if search else all_headers

grid_cols = st.columns(4)
selected = set(st.session_state.selected_headers)

for i, h in enumerate(filtered_headers):
    key = hdr_key(h)
    if key not in st.session_state:
        st.session_state[key] = (h in selected)
    grid_cols[i % 4].checkbox(h, key=key)

selected_headers = set()
for h in all_headers:
    if st.session_state.get(hdr_key(h), False):
        selected_headers.add(h)

st.session_state.selected_headers = selected_headers

st.divider()

btn_col1, btn_col2 = st.columns([1, 3])
with btn_col1:
    render_clicked = st.button("🚀 Xuất File", type="primary", use_container_width=True)
with btn_col2:
    st.caption("Sau khi chọn/bỏ cột xong, bấm nút để bắt đầu render PNG và tải ZIP.")

if not render_clicked:
    st.info("🟦 Chọn cột xong thì bấm **🚀 Xuất File** để bắt đầu.")
    st.stop()

if not selected_headers:
    st.error("Bạn đang bỏ chọn hết cột. Hãy tick ít nhất 1 cột để xuất ảnh.")
    st.stop()


# =========================
# PREPARE HTML LIST
# =========================
progress = st.progress(0.0)
status = st.empty()

cfg_css = build_css(watermark_size, stamp_size)

to_render = []
sheetnames = wb.sheetnames
total_sheets = len(sheetnames)

prepared = 0
skipped = 0

for idx_sheet, sheet_name in enumerate(sheetnames, start=1):
    status.info(f"Đang chuẩn bị sheet {idx_sheet}/{total_sheets}: **{sheet_name}**")
    ws = wb[sheet_name]

    data = []
    hl = []
    max_cols = 0

    for row in ws.iter_rows():
        row_vals = []
        row_hls = []
        for cell in row:
            row_vals.append(cell.value)
            row_hls.append(detect_highlight(cell))
        max_cols = max(max_cols, len(row_vals))
        data.append(row_vals)
        hl.append(row_hls)

    if len(data) < 2:
        skipped += 1
        progress.progress(idx_sheet / max(total_sheets, 1))
        continue

    for r in range(len(data)):
        if len(data[r]) < max_cols:
            data[r].extend([None] * (max_cols - len(data[r])))
        if len(hl[r]) < max_cols:
            hl[r].extend([False] * (max_cols - len(hl[r])))

    keep_cols = []
    for j in range(max_cols):
        if any(not is_empty_value(data[i][j]) for i in range(1, len(data))):
            keep_cols.append(j)

    if not keep_cols:
        skipped += 1
        progress.progress(idx_sheet / max(total_sheets, 1))
        continue

    headers_full = [normalize_header(data[0][j]) for j in keep_cols]

    keep_cols2 = []
    headers2 = []
    for j, hname in zip(keep_cols, headers_full):
        if hname and (hname in selected_headers):
            keep_cols2.append(j)
            headers2.append(hname)

    if not keep_cols2:
        skipped += 1
        progress.progress(idx_sheet / max(total_sheets, 1))
        continue

    body_rows_raw = [[data[i][j] for j in keep_cols2] for i in range(1, len(data))]
    hl_rows = [[hl[i][j] for j in keep_cols2] for i in range(1, len(hl))]
# ✅ FIX: tự tính và điền hàng TỔNG cho mọi cột số (ví dụ: T/ca ngày, T/ca đêm...)
fill_total_row_if_missing(body_rows_raw, hl_rows, headers2, force_overwrite=False)
    # Detect cột 100/130 dựa trên header GỐC (không bị rename)
    col_100 = None
    col_130 = None
    for i, h in enumerate(headers2):
        if col_100 is None and is_col_100(h):
            col_100 = i
        if col_130 is None and is_col_130(h):
            col_130 = i

    # Header hiển thị khi xuất ảnh: apply rename mapping
    headers_display = [rename_header_for_output(h) for h in headers2]
    headers_out = (["STT"] + headers_display) if add_stt else headers_display

    rows_for_html = []
    stt_counter = 0

    for r_idx, row in enumerate(body_rows_raw):
        is_total = is_total_row(row)

        stt_val = ""
        if add_stt and (not is_total):
            stt_counter += 1
            stt_val = str(stt_counter)

        low_bg_cols = set()
        low_text_cols = set()

        if (not is_total) and (col_100 is not None) and (col_130 is not None):
            raw100 = row[col_100]
            raw130 = row[col_130]
            has_data = (not is_empty_value(raw100)) or (not is_empty_value(raw130))
            if has_data:
                v100 = parse_float(raw100)
                v130 = parse_float(raw130)
                if (v100 + v130) < float(threshold_hours):
                    low_bg_cols.add(col_100)
                    low_bg_cols.add(col_130)

        elif (not is_total) and (col_100 is not None) and (col_130 is None):
            raw100 = row[col_100]
            if not is_empty_value(raw100):
                v100 = parse_float(raw100)
                if v100 < float(threshold_hours):
                    low_text_cols.add(col_100)

        cells = []

        if add_stt:
            cells.append({
                "value_html": html.escape(stt_val),
                "highlight": False,
                "lowhour": False,
                "lowtext": False,
            })

        for c_idx, val in enumerate(row):
            # Format theo header GỐC để đảm bảo nhận diện "giờ vào/ra", "ngày" không bị sai khi rename
            formatted = format_cell(val, headers2[c_idx])
            highlight = bool(hl_rows[r_idx][c_idx]) if r_idx < len(hl_rows) and c_idx < len(hl_rows[r_idx]) else False
            cells.append({
                "value_html": html.escape(formatted),
                "highlight": highlight,
                "lowhour": (c_idx in low_bg_cols),
                "lowtext": (c_idx in low_text_cols),
            })

        rows_for_html.append({"is_total": is_total, "cells": cells})

    base_name = safe_sheet_filename(sheet_name)
    numbered_name = f"{prepared + 1}_{base_name}.png"
    fname = numbered_name

    if stamp_mode == "Tên sheet":
        stamp_text = base_name
    elif stamp_mode == "Tên file ảnh":
        stamp_text = fname  # dùng tên đã đánh số
    else:
        stamp_text = (stamp_custom or "").strip()

    html_doc = build_html(
        sheet_name=sheet_name,
        headers=headers_out,
        rows=rows_for_html,
        stamp_text=stamp_text,
        cfg_css=cfg_css,
        watermark=watermark_text,
        show_wm=show_watermark,
        show_st=show_stamp,
    )

    to_render.append((fname, html_doc))
    prepared += 1
    progress.progress(idx_sheet / max(total_sheets, 1))

status.info("Đang xuất file ảnh (Chromium)...")

if not to_render:
    st.warning("Không có sheet nào có cột được chọn để xuất ảnh.")
    st.stop()


# =========================
# RENDER + ZIP + RAM + COUNT
# =========================
render_bar = st.progress(0.0)
render_text = st.empty()

def progress_cb(done, total, current_name):
    global _mem_peak_mb

    done = max(0, min(done, total))
    pct = 0.0 if total == 0 else done / total

    render_bar.progress(pct)
    render_text.info(f"Render {done}/{total}: **{current_name}**")

    # update RAM theo nhịp (đỡ tốn thời gian)
    if (done == 0) or (done == total) or (done % 3 == 0):
        cur_mb, peak_mb, sys_used_pct, sys_avail_mb, sys_total_mb = _read_ram()
        if cur_mb is not None:
            _mem_peak_mb = max(_mem_peak_mb, cur_mb)
        if peak_mb is not None:
            _mem_peak_mb = max(_mem_peak_mb, peak_mb)
        _render_mem_ui(cur_mb, _mem_peak_mb, sys_used_pct, sys_avail_mb, sys_total_mb)

# ZIP tối ưu RAM: dùng spooled file (nhỏ thì ở RAM, lớn sẽ tự đổ ra disk)
spool = SpooledTemporaryFile(max_size=64 * 1024 * 1024)  # 64MB
exported_count = 0

try:
    with zipfile.ZipFile(spool, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as z:
        exported_count = run_async(
            render_html_list_to_zip(
                to_render,
                z,
                progress_cb=progress_cb,
                dsf=device_scale,
                auto_fit=auto_fit_width,
                min_w=int(min_width),
                max_w=int(max_width),
                wait_ms=int(wait_ms),
            )
        )
except Exception as e:
    st.error(
        "Lỗi render Chromium.\n\n"
        f"Chi tiết: {e}"
    )
    st.stop()

render_bar.empty()
render_text.empty()

# đọc bytes zip để download
spool.seek(0)
zip_bytes = spool.read()
zip_mb = len(zip_bytes) / (1024 * 1024)

status.success(
    f"✅ Xong! Chuẩn bị {prepared}/{total_sheets} sheet (bỏ qua {skipped}), "
    f"đã xuất **{exported_count}** ảnh. ZIP ~ **{zip_mb:.2f} MB**"
)

st.download_button(
    "📥 Tải ZIP ảnh PNG",
    data=zip_bytes,
    file_name="bang_cong_png.zip",
    mime="application/zip",
)
